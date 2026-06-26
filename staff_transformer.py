"""
staff_transformer.py
--------------------
All transformation, cleaning, validation, date handling, shift derivation,
and template writing logic for the ArcGIS Special Event Staff Converter.

Separating all business logic from the UI (app.py) keeps this module
independently testable and reusable without Streamlit.

Main public functions:
    load_source_file()                  -- parse CSV or XLSX into raw DataFrame
    detect_source_format()              -- identify workup vs. staffing format
    find_current_staffing_header_row()  -- locate real column header in title-row CSVs
    reparse_staffing_from_raw()         -- reinterpret raw DataFrame with correct header
    transform_special_event_workup()    -- map SpecialEventWorkupforGIS columns to ArcGIS
    transform_current_staffing_report() -- map CurrentStaffingReport columns to ArcGIS
    validate_output()                   -- check required fields and ArcGIS constraints
    write_to_template()                 -- write output DataFrame into ArcGIS Excel template
"""

import re          # used by clean_text() to collapse multiple whitespace characters
import io          # used by write_to_template() to return an in-memory BytesIO buffer
import copy        # used by write_to_template() to deep-copy openpyxl cell formatting objects
import warnings    # reserved for future suppression of pandas/openpyxl deprecation warnings
from datetime import datetime, timedelta  # datetime arithmetic for UTC-to-local offset and shift math
from typing import Optional, Tuple        # type hints: Optional for nullable returns, Tuple for multi-value returns

import pandas as pd                          # DataFrame creation, row iteration, concat
from openpyxl import load_workbook           # read/write Excel workbooks while preserving all sheets and formatting
from openpyxl.utils import get_column_letter # convert 1-based column index to letter (available for future column sizing)
from dateutil import parser as dateutil_parser  # robust date-string parsing; handles many formats pandas cannot


# ===========================================================================
# ARCGIS OUTPUT SCHEMA
# ===========================================================================
# The ArcGIS Special Event Solution batch upload template requires EXACTLY
# these 20 column names in EXACTLY this order in the "Staff List" sheet.
# Any missing or misnamed column will cause ArcGIS to ignore or reject that field.
# This list drives all DataFrame construction and template writing throughout
# this module -- it is the single source of truth for output column order.
# ===========================================================================
ARCGIS_COLUMNS = [
    "unitid",         # unique identifier for the deployed unit (e.g., employee ID)
    "unitshift",      # human-readable shift label (e.g., "Days - 8 hr shift")
    "unitloc",        # physical deployment location -- typically blank, managed in ArcGIS
    "unittype",       # vehicle or assignment type (e.g., "Vehicle", "Traffic Control")
    "unitsquad",      # squad or unit number (e.g., "A123")
    "unitradio",      # radio call number assigned to the unit
    "unitduties",     # specific duties assigned to the unit
    "payrollnum",     # payroll or employee number -- often same value as unitid
    "staffrank",      # officer rank -- MUST match ArcGIS coded domain values exactly
    "staffname",      # full name in "LastName, FirstName" format
    "staffphone",     # cell phone number
    "staffemail",     # email address
    "staffskills",    # special skills (bilingual, K9, motorcycle, etc.)
    "staffpay",       # pay code or type
    "staffstatus",    # active status (e.g., "On Duty", "Off Duty")
    "staffduty",      # duty assignment or division
    "staffagency",    # employing agency -- always "HPD" for Houston Police Department
    "unitshiftstart", # shift start datetime -- written as Excel datetime with UTC offset applied
    "unitshiftend",   # shift end datetime -- written as Excel datetime with UTC offset applied
    "eventstatus",    # event status coded value (e.g., "Event Active")
]


# ===========================================================================
# COLUMN MAPPINGS -- SpecialEventWorkupforGIS.csv -> ArcGIS template
# ===========================================================================
# Keys   = ArcGIS output column names (from ARCGIS_COLUMNS above)
# Values = source CSV column names, or None when the field is always blank
#          or always set to a constant default value.
# This map documents the field-level relationship but is NOT iterated directly
# in the transform functions -- each field is handled explicitly so that
# special logic (rank normalization, datetime parsing, derived fields) can be
# applied cleanly per field.
# ===========================================================================
WORKUP_COLUMN_MAP = {
    "unitid":         "UnitId",
    "unitshift":      "UnitShift",     # if blank in source, derived from ShiftStart/ShiftEnd
    "unitloc":        None,            # always blank -- location is managed inside ArcGIS
    "unittype":       "UnitType",      # default: "Vehicle" (configurable in sidebar)
    "unitsquad":      "UnitSquad",
    "unitradio":      "UnitRadio",
    "unitduties":     "UnitDuties",
    "payrollnum":     "Payroll",
    "staffrank":      "StaffRank",     # normalized via normalize_staff_rank()
    "staffname":      "StaffName",
    "staffphone":     "StaffPhone",
    "staffemail":     "Staffemail",    # lowercase 'e' in source -- matches actual CSV header spelling
    "staffskills":    "StaffSkills",
    "staffpay":       "StaffPay",
    "staffstatus":    "StaffStatus",   # default: "On Duty"
    "staffduty":      "StaffDuty",
    "staffagency":    "StaffAgency",   # always forced to "HPD" regardless of source value
    "unitshiftstart": "ShiftStart",    # UTC datetime string; +offset_hours applied
    "unitshiftend":   "ShiftEnd",      # UTC datetime string; +offset_hours applied
    "eventstatus":    None,            # always "Event Active" -- no source column
}

# Subset of datetime fields from the workup source that need UTC offset applied.
# Stored separately to avoid re-checking all 20 columns in every row loop.
WORKUP_DATETIME_FIELDS = {
    "unitshiftstart": "ShiftStart",
    "unitshiftend":   "ShiftEnd",
}


# ===========================================================================
# COLUMN MAPPINGS -- CurrentStaffingReport.csv -> ArcGIS template
# ===========================================================================
# The staffing report uses different column names and requires derived fields:
#   staffname  = "{LastName}, {FirstName}"  (two source columns combined)
#   unitshift  = derived from shift/shiftStart/ShiftEnd datetime columns
#   unitid     = EmpID (same source column as payrollnum)
# ===========================================================================
STAFFING_COLUMN_MAP = {
    "unitid":         "EmpID",           # employee ID doubles as the unit identifier
    "unitshift":      None,              # derived from shift/shiftStart/ShiftEnd
    "unitloc":        None,              # always blank
    "unittype":       None,              # always "Traffic Control" (configurable default)
    "unitsquad":      "UnitNo",          # unit/squad number from the report
    "unitradio":      "RadioCallNumber", # radio call sign
    "unitduties":     None,              # not present in staffing report -- always blank
    "payrollnum":     "EmpID",           # same source column as unitid
    "staffrank":      "RankDescription", # normalized via normalize_staff_rank()
    "staffname":      None,              # derived: "{LastName}, {FirstName}"
    "staffphone":     "CellPhone",
    "staffemail":     None,              # not in staffing report
    "staffskills":    None,              # not in staffing report
    "staffpay":       None,              # not in staffing report
    "staffstatus":    None,              # always "On Duty" (default)
    "staffduty":      "Division",        # officer's assigned division
    "staffagency":    None,              # always "HPD"
    "unitshiftstart": "shiftStart",      # note lowercase 's' -- must match source column exactly
    "unitshiftend":   "ShiftEnd",        # capital 'S' -- matches source column exactly
    "eventstatus":    None,              # always "Event Active"
}

# Datetime fields in the staffing report that need UTC offset applied.
STAFFING_DATETIME_FIELDS = {
    "unitshiftstart": "shiftStart",
    "unitshiftend":   "ShiftEnd",
}


# ===========================================================================
# STAFFING HEADER IDENTIFIERS
# ===========================================================================
# CurrentStaffingReport.csv starts with report title rows before the real
# column header. Example structure:
#
#   Row 0: "Textbox23"           (report title -- only 1 field)
#   Row 1: "Total Records: 142"  (record count metadata)
#   Row 2: ""                    (blank separator)
#   Row 3: "Division, RankDescription, LastName, FirstName, ..."  (REAL HEADER)
#   Row 4+: data rows
#
# These strings appear as CELL VALUES in the real header row. We scan each row
# looking for >= 5 of these identifiers to locate the actual column header row.
# Requiring 5 matches guards against false positives on data rows.
# ===========================================================================
STAFFING_HEADER_IDENTIFIERS = {
    "Division", "RankDescription", "LastName", "FirstName",
    "EmpID", "RadioCallNumber", "UnitNo", "shiftStart", "ShiftEnd",
}

# ===========================================================================
# TEXT ID COLUMNS
# ===========================================================================
# Column names whose values must be preserved as plain strings and never
# converted to float. For example, EmpID "00142" must not become 142.0.
# Actual string enforcement is handled by dtype=str in load_source_file();
# this set documents the intent and can be used for future post-read casting.
# ===========================================================================
TEXT_ID_COLUMNS = {
    "empid", "payroll", "unitid", "unitno", "radiocallnumber", "unitradio",
}


# ===========================================================================
# UTILITY HELPERS
# ===========================================================================

def clean_text(value) -> str:
    """
    Normalize a cell value to a clean, trimmed string.

    Handles pd.NA, None, float NaN, and regular strings uniformly so that
    callers never need to type-check before string operations.

    Collapses multiple consecutive spaces into one so that ArcGIS never
    receives values like "SMITH  JOHN" that would fail a name lookup.

    Returns "" for any null-like input -- ArcGIS prefers empty strings
    over "None" or "nan" as literal cell text.
    """
    if pd.isna(value) or value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r" {2,}", " ", s)  # collapse runs of 2+ spaces into one
    return s


def _series_clean(series: pd.Series) -> pd.Series:
    """
    Vectorized version of clean_text() for an entire Series at once.
    More efficient than applying clean_text() row-by-row when bulk column
    cleaning is needed (e.g., cleaning all rows of a single column).
    """
    return series.astype(str).str.strip().str.replace(r" {2,}", " ", regex=True)


def _case_insensitive_col(df: pd.DataFrame, name: str) -> Optional[str]:
    """
    Return the actual column name in df that matches `name` case-insensitively.

    WHY THIS IS NEEDED:
    1. Source CSV column names vary in capitalization across export versions
       (e.g., "EmpID" vs "empid" vs "EMPID").
    2. When load_source_file() reads a CSV before the header row is identified,
       pandas assigns integer column names (0, 1, 2, ...). The str() cast
       prevents AttributeError when those integers reach the .strip() call.

    Returns the matching column name string, or None if not found.
    Callers treat None as "column absent from this file."
    """
    for col in df.columns:
        if str(col).strip().lower() == name.lower():
            return col
    return None


# ===========================================================================
# STAFF RANK NORMALIZATION
# ===========================================================================

# ---------------------------------------------------------------------------
# APPROVED RANKS -- ArcGIS coded domain values
# ---------------------------------------------------------------------------
# The ArcGIS Special Event Solution enforces a CLOSED domain for staffrank.
# Values must match these strings EXACTLY -- capitalization, punctuation,
# and spacing are all significant:
#   - "Senior Police Officer" is mixed case, NOT all-caps
#   - "POLICE OFFICER,PROBATIONARY" has NO space after the comma
#   - "Civilian" is title case, NOT all-caps
# Any value outside this list will fail ArcGIS domain validation on import.
# ---------------------------------------------------------------------------
APPROVED_RANKS = [
    "POLICE CHIEF",
    "EXECUTIVE CHIEF",
    "EXECUTIVE ASSISTANT POLICE CHIEF",
    "ASSISTANT POLICE CHIEF",
    "POLICE CAPTAIN",
    "POLICE LIEUTENANT",
    "POLICE SERGEANT",
    "Senior Police Officer",        # intentional mixed case -- must match ArcGIS exactly
    "POLICE OFFICER",
    "POLICE OFFICER,PROBATIONARY",  # no space after comma -- must match ArcGIS exactly
    "Civilian",                     # intentional title case -- must match ArcGIS exactly
    "OLEA",                         # Outside Law Enforcement Agency
]

# Fast lowercase lookup dict: "police chief" -> "POLICE CHIEF"
# Built from APPROVED_RANKS so it auto-updates when the approved list changes.
_RANK_CANONICAL = {r.lower(): r for r in APPROVED_RANKS}

# ---------------------------------------------------------------------------
# RANK ALIASES
# ---------------------------------------------------------------------------
# Maps common HPD source abbreviations and alternate spellings to the exact
# approved ArcGIS rank value. All keys are lowercase so the lookup is
# case-insensitive. Aliases are checked BEFORE the canonical lookup so that
# common shorthand (e.g., "SGT" -> "POLICE SERGEANT") is matched first,
# avoiding ambiguity with partial canonical matches.
# ---------------------------------------------------------------------------
_RANK_ALIASES: dict = {
    # Police Chief
    "chief of police":                      "POLICE CHIEF",
    "police chief":                         "POLICE CHIEF",
    "chief":                                "POLICE CHIEF",

    # Executive Chief
    "executive chief":                      "EXECUTIVE CHIEF",

    # Executive Assistant Police Chief
    "executive ac":                         "EXECUTIVE ASSISTANT POLICE CHIEF",
    "executive assistant chief":            "EXECUTIVE ASSISTANT POLICE CHIEF",
    "executive assistant police chief":     "EXECUTIVE ASSISTANT POLICE CHIEF",
    "exec ac":                              "EXECUTIVE ASSISTANT POLICE CHIEF",

    # Assistant Police Chief
    "assistant chief":                      "ASSISTANT POLICE CHIEF",
    "assistant police chief":               "ASSISTANT POLICE CHIEF",
    "asst chief":                           "ASSISTANT POLICE CHIEF",

    # Police Captain
    "captain":                              "POLICE CAPTAIN",
    "police captain":                       "POLICE CAPTAIN",
    "capt":                                 "POLICE CAPTAIN",

    # Police Lieutenant
    "lieutenant":                           "POLICE LIEUTENANT",
    "police lieutenant":                    "POLICE LIEUTENANT",
    "lt":                                   "POLICE LIEUTENANT",

    # Police Sergeant
    "sergeant":                             "POLICE SERGEANT",
    "police sergeant":                      "POLICE SERGEANT",
    "sgt":                                  "POLICE SERGEANT",

    # Senior Police Officer
    "senior police officer":                "Senior Police Officer",
    "sr police officer":                    "Senior Police Officer",
    "senior officer":                       "Senior Police Officer",

    # Police Officer
    "police officer":                       "POLICE OFFICER",
    "officer":                              "POLICE OFFICER",
    "po":                                   "POLICE OFFICER",
    "police ofc":                           "POLICE OFFICER",

    # Police Officer Probationary
    "police officer probationary":          "POLICE OFFICER,PROBATIONARY",
    "police officer, probationary":         "POLICE OFFICER,PROBATIONARY",
    "probationary police officer":          "POLICE OFFICER,PROBATIONARY",
    "po probationary":                      "POLICE OFFICER,PROBATIONARY",

    # Civilian
    "civilian":                             "Civilian",
    "hpd civilian":                         "Civilian",

    # Outside Law Enforcement Agency
    "outside leo":                          "OLEA",
    "olea":                                 "OLEA",
}


def normalize_staff_rank(value: str) -> Tuple[str, bool]:
    """
    Map a source rank string to the exact ArcGIS approved staffrank coded value.

    Two-stage lookup strategy:
      Stage 1 -- check _RANK_ALIASES first.
                 Covers HPD abbreviations (SGT, LT, PO) and common alternate
                 spellings that differ from the ArcGIS canonical form.
      Stage 2 -- fall through to _RANK_CANONICAL.
                 Handles the case where the source already uses the correct
                 ArcGIS value but with different capitalization.

    Returns:
        (normalized_rank, matched)
        - normalized_rank: the approved ArcGIS string if found;
                           otherwise the cleaned original (row is not dropped)
        - matched: True if a known mapping was found; False signals the
                   caller to emit a warning so the analyst can review

    Blank input returns ("", True): blank rank is valid for some row types
    and should NOT generate a warning -- we distinguish "missing rank" from
    "unrecognized rank" at the validation layer.
    """
    cleaned = clean_text(value)
    if not cleaned:
        return "", True   # blank -- no warning needed

    key = cleaned.lower()

    # Stage 1: alias table (HPD shorthand -> approved ArcGIS value)
    if key in _RANK_ALIASES:
        return _RANK_ALIASES[key], True

    # Stage 2: canonical table (case-insensitive exact match against approved list)
    if key in _RANK_CANONICAL:
        return _RANK_CANONICAL[key], True

    # No match -- return original string unchanged so the row is not silently blanked.
    # The caller will warn the analyst; the row will still appear in the output.
    return cleaned, False


def _get(df: pd.DataFrame, row: pd.Series, source_col: str) -> str:
    """
    Case-insensitive single-cell getter that always returns a clean string.

    Looks up source_col in df.columns case-insensitively (via _case_insensitive_col),
    then returns the cleaned string value from the given row.

    Returns "" if the column is not present in df -- callers can treat ""
    the same as a blank source cell without additional None checks.
    """
    actual = _case_insensitive_col(df, source_col)
    if actual is None:
        return ""
    return clean_text(row.get(actual, ""))


# ===========================================================================
# FILE LOADING
# ===========================================================================

def load_source_file(file_obj) -> Tuple[pd.DataFrame, str]:
    """
    Load a CSV or XLSX staffing file into a raw DataFrame with all values as strings.

    Returns: (dataframe, file_extension)

    WHY THE PYTHON csv MODULE INSTEAD OF PANDAS FOR CSV FILES:
    CurrentStaffingReport.csv has variable-width rows:

        Row 1: "Textbox23"                          <- 1 field (report title)
        Row 2: "Total Records: 142"                 <- 1 field (metadata)
        Row 3: ""                                   <- blank
        Row 4: "Division,RankDescription,LastName,..."  <- 13 fields (real header)
        Row 5+: data rows with 13 fields each

    Pandas' C parser locks in the expected column count from the FIRST row.
    When row 1 has 1 field and row 5 has 13, pandas raises:
        "Expected 1 fields in line 5, saw 13"

    Using engine='python' does not help -- it raises the same error.
    Adding on_bad_lines='skip' silently drops ALL data rows (they all look
    "bad" relative to the 1-field first row).

    Solution: Python's built-in csv module reads each row independently,
    producing lists of different lengths. We then pad every row to the
    maximum field count (with empty strings) so pandas can build a uniform
    DataFrame. The resulting DataFrame has integer column names (0, 1, 2, ...)
    which are resolved to real column names by the header-row detection step.

    For XLSX files, pandas read_excel handles parsing normally; the
    variable-width issue does not apply.

    Args:
        file_obj: file-like object with a .name attribute (Streamlit UploadedFile
                  or io.BytesIO with .name set). Must be positioned at byte 0.

    Returns:
        (DataFrame with dtype=str, lowercase extension string)

    Raises:
        ValueError: if the file is empty or has an unsupported extension.
    """
    import csv as _csv  # local import to keep the module namespace clean
    import io as _io    # local import -- mirrors the stdlib io already imported above

    # Determine file type from extension to route CSV vs XLSX handling
    name = getattr(file_obj, "name", "")
    ext  = name.rsplit(".", 1)[-1].lower() if "." in name else ""

    if ext == "csv":
        # Read the entire file into memory as a string.
        # errors="replace" prevents UnicodeDecodeError on files that contain
        # non-UTF-8 characters (common in legacy police staffing export tools).
        raw = file_obj.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8", errors="replace")

        # Parse all rows using Python's csv reader.
        # Handles quoted fields, commas inside values, and variable row widths correctly.
        reader = _csv.reader(_io.StringIO(raw))
        rows   = list(reader)

        if not rows:
            raise ValueError("CSV file is empty.")

        # Find the widest row to determine the DataFrame column count.
        # Title rows have fewer fields; data rows have more.
        # Padding short rows with "" ensures all rows are the same width.
        max_cols = max(len(r) for r in rows)
        padded   = [r + [""] * (max_cols - len(r)) for r in rows]

        # All columns get integer names (0, 1, 2, ...) at this stage.
        # detect_source_format() or find_current_staffing_header_row() will
        # locate the real header row; reparse_staffing_from_raw() will rename
        # the columns to their actual string names.
        df = pd.DataFrame(padded, dtype=str)

    elif ext in ("xlsx", "xls"):
        # pandas handles Excel files without the variable-width issue.
        # keep_default_na=False prevents pandas from silently converting
        # the text "NA" or "N/A" into NaN float values.
        df = pd.read_excel(file_obj, dtype=str, keep_default_na=False)
    else:
        raise ValueError(f"Unsupported file type: .{ext}. Expecting .csv or .xlsx")

    return df, ext


# ===========================================================================
# FORMAT DETECTION
# ===========================================================================

def detect_source_format(df: pd.DataFrame) -> str:
    """
    Identify whether df came from SpecialEventWorkupforGIS or CurrentStaffingReport.

    Strategy: score each known format by counting how many of its distinctive
    column names appear in df.columns (case-insensitive). The format with
    the higher score wins, provided it meets a minimum threshold of 3 matches.

    Returns: 'workup', 'staffing', or 'unknown'

    'unknown' is returned when:
    - Neither format reaches the 3-column threshold.
    - This is the expected result when the DataFrame still has integer column
      names (0, 1, 2, ...) from a title-row CSV -- at that point the columns
      ARE the title rows, not the real headers.

    When 'unknown' is returned, the caller should:
      1. Call find_current_staffing_header_row(raw_df) to locate the real header row.
      2. Call reparse_staffing_from_raw(raw_df, idx) to rebuild with proper column names.
      3. Call detect_source_format() again on the reparsed DataFrame.
    """
    # Cast all column names to strings BEFORE calling .strip() or .lower().
    # When loaded without a header row, pandas assigns integer column indices
    # (0, 1, 2, ...) which raise AttributeError if you call .strip() directly.
    cols_lower = {str(c).strip().lower() for c in df.columns}

    # Distinctive column names for each format -- chosen to appear in one
    # format but not the other to minimize false positives.
    workup_signals   = {"unitid", "staffrank", "staffname", "shiftstart", "shiftend"}
    staffing_signals = {"rankdescription", "lastname", "firstname", "empid", "shiftstart", "shiftend"}

    workup_score   = len(workup_signals   & cols_lower)
    staffing_score = len(staffing_signals & cols_lower)

    # Require a clear winner AND a minimum signal count
    if workup_score > staffing_score and workup_score >= 3:
        return "workup"
    if staffing_score > workup_score and staffing_score >= 3:
        return "staffing"
    return "unknown"


# ===========================================================================
# CurrentStaffingReport: HEADER ROW DETECTION AND REPARSING
# ===========================================================================

def find_current_staffing_header_row(raw_df: pd.DataFrame) -> Optional[int]:
    """
    Scan a raw DataFrame row by row to find the real column header row,
    by looking for known column name strings as CELL VALUES.

    WHY WE SCAN VALUES (not column names):
    At this point raw_df has integer column names (0, 1, 2, ...) because
    the CSV was loaded with no header. The actual column names exist as
    data values in one of the rows. We scan each row's values to find
    which row contains the real column headers.

    Match threshold: >= 5 of STAFFING_HEADER_IDENTIFIERS must appear as
    cell values in a single row. This prevents false matches on data rows
    that might coincidentally contain one or two of the identifier strings.

    Returns:
        0-based integer index into raw_df for the header row, or None
        if no header row is found (unexpected file structure).
    """
    for idx, row in raw_df.iterrows():
        cell_values = {str(v).strip() for v in row.values}
        if len(STAFFING_HEADER_IDENTIFIERS & cell_values) >= 5:
            return idx
    return None


def reload_staffing_with_real_header(file_obj, header_row_index: int) -> pd.DataFrame:
    """
    Re-read the CSV using `header_row_index` as the column header row.

    LEGACY -- kept for reference but NOT used as the primary path.
    Replaced by reparse_staffing_from_raw() which avoids two problems:
      1. File re-reads require seek(0) which is unreliable on Streamlit Cloud.
      2. pandas' header= parameter counts file lines from 0, while our
         detected index is a raw_df row index -- these can differ when the
         CSV has multiline fields or rows with only commas.

    Args:
        file_obj:          file-like object seeked back to position 0
        header_row_index:  0-based file line index of the header row
    """
    file_obj.seek(0)
    df = pd.read_csv(
        file_obj,
        dtype=str,
        keep_default_na=False,
        header=header_row_index,
    )
    return df


def reparse_staffing_from_raw(raw_df: pd.DataFrame, header_row_idx: int) -> pd.DataFrame:
    """
    Reinterpret an already-loaded raw DataFrame by treating one row as the
    column header and all following rows as data.

    WHY THIS INSTEAD OF reload_staffing_with_real_header():
      1. No file re-read. Streamlit's UploadedFile.seek(0) is not reliable
         after the file has already been read -- position may not reset on
         all platforms. We buffer bytes upfront in app.py (_make_buf()) and
         avoid any re-read here entirely.
      2. No off-by-one bugs. pandas' read_csv(header=N) counts FILE LINES
         from 0. Our raw_df row index is a DATAFRAME row index -- these
         diverge when title rows contain commas or when the CSV has a BOM.
         Slicing the DataFrame directly is unambiguous.

    Args:
        raw_df:          full raw DataFrame including title rows (integer column names)
        header_row_idx:  0-based raw_df index returned by find_current_staffing_header_row()

    Returns:
        New DataFrame with column names from row `header_row_idx` and data
        starting from the next row. Index reset to 0.
    """
    # Extract column name strings from the identified header row
    new_cols = [str(v).strip() for v in raw_df.iloc[header_row_idx].tolist()]

    # Slice everything after the header row -- these are the actual data rows
    data = raw_df.iloc[header_row_idx + 1:].copy()
    data.columns = new_cols

    # Reset index so downstream row iteration starts at 0 (cleaner row numbers in warnings)
    data = data.reset_index(drop=True)
    return data


# ===========================================================================
# DATETIME PARSING AND UTC OFFSET
# ===========================================================================

# Excel/ArcGIS datetime number format applied to unitshiftstart and unitshiftend cells.
# ArcGIS expects this specific format when reading dates from the Staff List sheet.
_EXCEL_DATE_FORMAT    = "yyyy/m/dd h:mm:ss AM/PM"
_OPENPYXL_DATE_FORMAT = "yyyy/m/dd h:mm:ss AM/PM"  # same value -- kept explicit for readability


def parse_and_offset_datetime(
    value: str,
    offset_hours: float = 5,
) -> Tuple[Optional[datetime], bool]:
    """
    Parse a date/time string and apply a UTC-to-local hour offset.

    WHY THE OFFSET:
    Source staffing reports store shift times in UTC. ArcGIS and the HPD
    event dashboard display times in Central time. The default +5 converts
    UTC to Central Daylight Time (CDT = UTC-5). For Central Standard Time
    (CST = UTC-6) use +6. This is configurable via the Streamlit sidebar
    so the analyst can adjust seasonally or if the source system changes.

    WHY python-dateutil INSTEAD OF strptime():
    Source datetime strings vary in format across export versions:
        "2024-06-15 08:00:00"
        "6/15/2024 8:00 AM"
        "2024-06-15T08:00:00Z"
    dateutil_parser.parse() handles all of these without a format string.
    strptime() would require a known fixed format that could break on
    any future export format change.

    Args:
        value:        raw date/time string from the source CSV
        offset_hours: hours to add after parsing (default 5 for UTC->CDT)

    Returns:
        (datetime_object, True)  on success -- caller writes datetime to Excel cell
        (None, False)            on failure -- caller keeps raw string and emits warning
    """
    if not value or value.strip() == "":
        return None, False  # blank field -- not an error, no warning needed

    try:
        # dayfirst=False: treat ambiguous dates as MM/DD/YYYY not DD/MM/YYYY
        dt = dateutil_parser.parse(value, dayfirst=False)
        dt_offset = dt + timedelta(hours=offset_hours)
        return dt_offset, True
    except Exception:
        return None, False


# ===========================================================================
# SHIFT LABEL DERIVATION
# ===========================================================================

def derive_unit_shift(start_str: str, end_str: str) -> str:
    """
    Compute a human-readable shift label from shift start and end time strings.

    Output format: "{Daypart} - {Duration label}"

    Examples:
        start=06:00, end=14:00  ->  "Days - 8 hr shift"
        start=16:00, end=02:00  ->  "Evenings - 10 hr shift"  (crosses midnight)
        start=22:00, end=10:00  ->  "Nights - 12 hr shift"    (crosses midnight)

    DAYPART RULES (based on start hour):
        05:00 - 14:59  -->  Days
        15:00 - 21:59  -->  Evenings
        22:00+, 00:00 - 04:59  -->  Nights

    DURATION BUCKETS (rounded from actual elapsed hours):
        <= 9 hours   -->  "8 hr shift"
        <= 11 hours  -->  "10 hr shift"
        > 11 hours   -->  "12 hr shift"

    MIDNIGHT CROSSING:
    If end_dt <= start_dt after parsing (e.g., end "06:00" < start "22:00"),
    we assume the shift crosses midnight and add 24 hours to end_dt.
    Without this, the duration would be negative and the daypart label wrong.

    Returns "" if either start or end cannot be parsed -- the caller
    will store an empty unitshift, which is valid in ArcGIS.
    """
    try:
        start_dt = dateutil_parser.parse(start_str, dayfirst=False)
        end_dt   = dateutil_parser.parse(end_str,   dayfirst=False)
    except Exception:
        return ""  # unparseable times -- return blank, not an error

    # Handle shifts that cross midnight
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)

    duration_hours = (end_dt - start_dt).total_seconds() / 3600

    # Map actual duration to nearest standard HPD shift length
    if duration_hours <= 9:
        hrs_label = "8 hr shift"
    elif duration_hours <= 11:
        hrs_label = "10 hr shift"
    else:
        hrs_label = "12 hr shift"

    # Classify daypart by the start hour
    start_hour = start_dt.hour
    if 5 <= start_hour <= 14:
        daypart = "Days"
    elif 15 <= start_hour <= 21:
        daypart = "Evenings"
    else:
        daypart = "Nights"

    return f"{daypart} - {hrs_label}"


# ===========================================================================
# TRANSFORM: SpecialEventWorkupforGIS.csv -> ArcGIS columns
# ===========================================================================

def transform_special_event_workup(
    df: pd.DataFrame,
    offset_hours: float = 5,
    default_unit_type: str = "Vehicle",
    default_unit_radio: str = "",
    default_unit_duties: str = "",
    default_staff_status: str = "On Duty",
    default_staff_agency: str = "HPD",
    default_event_status: str = "Event Active",
) -> Tuple[pd.DataFrame, list]:
    """
    Map SpecialEventWorkupforGIS.csv columns into the 20-column ArcGIS schema.

    Each row is processed individually (not vectorized) so that per-row warnings
    can reference the source row number. Row numbers are reported as i+2 to
    match what the analyst sees in Excel (1-based + 1 for the header row).

    Default values from the sidebar are applied when source fields are blank.
    staffagency is ALWAYS "HPD" regardless of sidebar or source -- business rule.

    Args:
        df:                   cleaned DataFrame with correct column names
        offset_hours:         hours to add to shift datetimes for UTC->local conversion
        default_unit_type:    used when UnitType column is blank (typically "Vehicle")
        default_staff_status: used when StaffStatus column is blank (typically "On Duty")
        default_staff_agency: accepted for API consistency but ignored; "HPD" is hardcoded
        default_event_status: written to all rows (typically "Event Active")

    Returns:
        (output_df, warnings_list)
        - output_df has ARCGIS_COLUMNS as columns, in order
        - warnings_list has one entry per data quality issue
    """
    warnings_list = []
    rows = []

    # Drop fully blank rows before processing.
    # Blank rows produce empty output rows that fail ArcGIS validation.
    # Replace "" with pd.NA first so dropna(how="all") works correctly.
    df = df.replace("", pd.NA).dropna(how="all").fillna("").reset_index(drop=True)

    for i, row in df.iterrows():
        out = {}  # accumulates all 20 ArcGIS field values for this row

        # unitid: direct source column mapping
        out["unitid"] = _get(df, row, "UnitId")

        # unitshift: use source value if present; otherwise derive from shift times.
        # Derived labels follow "Days/Evenings/Nights - N hr shift" pattern.
        unit_shift_raw = _get(df, row, "UnitShift")
        if unit_shift_raw:
            out["unitshift"] = unit_shift_raw
        else:
            ss = _get(df, row, "ShiftStart")
            se = _get(df, row, "ShiftEnd")
            out["unitshift"] = derive_unit_shift(ss, se)

        # unitloc: always blank -- deployment location is managed inside ArcGIS
        out["unitloc"] = ""

        # unittype: source value takes precedence; falls back to sidebar default
        ut = _get(df, row, "UnitType")
        out["unittype"] = ut if ut else default_unit_type

        # Direct column mappings requiring no transformation
        out["unitsquad"]  = _get(df, row, "UnitSquad")
        radio = _get(df, row, "UnitRadio")
        out["unitradio"]  = radio if radio else default_unit_radio
        duties = _get(df, row, "UnitDuties")
        out["unitduties"] = duties if duties else default_unit_duties
        out["payrollnum"] = _get(df, row, "Payroll")

        # staffrank: normalize source value to exact ArcGIS coded domain string.
        # Row is kept even if rank is not matched -- analyst sees warning and can correct.
        raw_rank = _get(df, row, "StaffRank")
        norm_rank, rank_matched = normalize_staff_rank(raw_rank)
        out["staffrank"] = norm_rank
        if not rank_matched:
            warnings_list.append(
                f"Row {i + 2}: Unknown staffrank '{raw_rank}' — kept as-is, verify against approved list."
            )

        out["staffname"]   = _get(df, row, "StaffName")
        out["staffphone"]  = _get(df, row, "StaffPhone")
        out["staffemail"]  = _get(df, row, "Staffemail")  # lowercase 'e' matches source header
        out["staffskills"] = _get(df, row, "StaffSkills")
        out["staffpay"]    = _get(df, row, "StaffPay")

        # staffstatus: use source if present, else sidebar default
        ss_val = _get(df, row, "StaffStatus")
        out["staffstatus"] = ss_val if ss_val else default_staff_status

        out["staffduty"] = _get(df, row, "StaffDuty")

        # staffagency: ALWAYS "HPD" -- hard business rule for HPD deployments.
        # The default_staff_agency argument is accepted for API consistency but not used.
        out["staffagency"] = "HPD"

        # unitshiftstart and unitshiftend: parse UTC datetime strings and apply offset.
        # On success, store a Python datetime object so openpyxl writes a proper Excel
        # datetime serial number. On failure, store the raw string so the cell is not
        # silently blank, and emit a warning for the analyst to investigate.
        for arcgis_col, src_col in WORKUP_DATETIME_FIELDS.items():
            raw_val = _get(df, row, src_col)
            dt_obj, ok = parse_and_offset_datetime(raw_val, offset_hours)
            if ok:
                out[arcgis_col] = dt_obj
            else:
                out[arcgis_col] = raw_val
                if raw_val:  # only warn when there was a non-blank value that failed to parse
                    warnings_list.append(
                        f"Row {i + 2}: Could not parse {arcgis_col} value '{raw_val}'"
                    )

        # eventstatus: constant for all rows
        out["eventstatus"] = default_event_status

        rows.append(out)

    # Enforce ARCGIS_COLUMNS order explicitly.
    # Being explicit here prevents any future dict insertion order surprise
    # from affecting the column order written to the Excel template.
    result = pd.DataFrame(rows, columns=ARCGIS_COLUMNS)
    return result, warnings_list


# ===========================================================================
# TRANSFORM: CurrentStaffingReport.csv -> ArcGIS columns
# ===========================================================================

def transform_current_staffing_report(
    df: pd.DataFrame,
    offset_hours: float = 5,
    default_unit_type: str = "Vehicle",
    default_unit_radio: str = "",
    default_unit_duties: str = "",
    default_staff_status: str = "On Duty",
    default_staff_agency: str = "HPD",
    default_event_status: str = "Event Active",
) -> Tuple[pd.DataFrame, list]:
    """
    Map CurrentStaffingReport.csv columns into the 20-column ArcGIS schema.

    Key differences from transform_special_event_workup():
    - staffname is DERIVED: "{LastName}, {FirstName}" -- not a single source column
    - unitid and payrollnum both map to EmpID (same source column, two output fields)
    - unitshift is derived from shift/shiftStart/ShiftEnd (no direct source column)
    - Many fields are blank in this source: unitloc, unitduties, staffemail,
      staffskills, staffpay -- these are always written as empty strings
    - default_unit_type defaults to "Traffic Control" (vs "Vehicle" for workup)

    Args:
        df:                   cleaned DataFrame with real column headers (after reparsing)
        offset_hours:         hours to add to shift datetimes
        default_unit_type:    applied to all rows (typically "Traffic Control")
        default_staff_status: applied to all rows (typically "On Duty")
        default_staff_agency: accepted for API consistency; "HPD" is hardcoded
        default_event_status: written to all rows (typically "Event Active")

    Returns:
        (output_df, warnings_list)
    """
    warnings_list = []
    rows = []

    # Strip fully blank rows before processing
    df = df.replace("", pd.NA).dropna(how="all").fillna("").reset_index(drop=True)

    for i, row in df.iterrows():
        out = {}

        emp_id = _get(df, row, "EmpID")  # used for both unitid and payrollnum

        # unitid and payrollnum: both populated from EmpID.
        # ArcGIS treats them as separate fields but HPD uses the same value for both.
        out["unitid"]    = emp_id
        out["unitloc"]   = ""                  # always blank -- not in staffing report
        out["unittype"]  = default_unit_type   # "Traffic Control" for all staffing rows
        out["unitsquad"] = _get(df, row, "UnitNo")
        radio = _get(df, row, "RadioCallNumber")
        out["unitradio"] = radio if radio else default_unit_radio
        out["unitduties"]= default_unit_duties  # not in staffing report; use sidebar default
        out["payrollnum"]= emp_id              # same as unitid

        # staffrank: RankDescription contains HPD rank names.
        # normalize_staff_rank() maps abbreviations (SGT, LT) and alternate
        # spellings to the exact ArcGIS coded domain value.
        raw_rank = _get(df, row, "RankDescription")
        norm_rank, rank_matched = normalize_staff_rank(raw_rank)
        out["staffrank"] = norm_rank
        if not rank_matched:
            warnings_list.append(
                f"Row {i + 2}: Unknown staffrank '{raw_rank}' — kept as-is, verify against approved list."
            )

        # staffname: ArcGIS expects "LastName, FirstName" format.
        # The staffing report provides these as separate columns.
        # Handle all combinations gracefully: both present, only one, or both blank.
        last  = _get(df, row, "LastName")
        first = _get(df, row, "FirstName")
        if last and first:
            full = f"{last}, {first}"   # standard HPD format: "SMITH, JOHN"
        elif last:
            full = last                 # only last name found
        elif first:
            full = first                # only first name found (unusual)
        else:
            full = ""                   # both blank -- triggers blank staffname warning in validate_output()
        out["staffname"] = full

        out["staffphone"]  = _get(df, row, "CellPhone")
        out["staffemail"]  = ""              # not available in staffing report
        out["staffskills"] = ""              # not available in staffing report
        out["staffpay"]    = ""              # not available in staffing report
        out["staffstatus"] = default_staff_status   # always "On Duty"
        out["staffduty"]   = _get(df, row, "Division")
        out["staffagency"] = "HPD"           # always hardcoded -- HPD business rule

        # unitshift: staffing report may have an explicit "shift" column or
        # we fall back to deriving it from the actual start/end datetime strings.
        shift_label = _get(df, row, "shift")   # optional explicit label (some exports include this)
        ss_raw = _get(df, row, "shiftStart")    # lowercase 's' -- must match source column exactly
        se_raw = _get(df, row, "ShiftEnd")

        if shift_label:
            out["unitshift"] = shift_label      # use explicit label if present
        else:
            out["unitshift"] = derive_unit_shift(ss_raw, se_raw)

        # unitshiftstart and unitshiftend: parse and apply UTC offset.
        for arcgis_col, src_col in STAFFING_DATETIME_FIELDS.items():
            raw_val = _get(df, row, src_col)
            dt_obj, ok = parse_and_offset_datetime(raw_val, offset_hours)
            if ok:
                out[arcgis_col] = dt_obj
            else:
                out[arcgis_col] = raw_val
                if raw_val:
                    warnings_list.append(
                        f"Row {i + 2}: Could not parse {arcgis_col} value '{raw_val}'"
                    )

        out["eventstatus"] = default_event_status

        rows.append(out)

    result = pd.DataFrame(rows, columns=ARCGIS_COLUMNS)
    return result, warnings_list


# ===========================================================================
# VALIDATION
# ===========================================================================

def validate_output(df: pd.DataFrame, source_df: pd.DataFrame = None) -> list:
    """
    Run data quality checks on the fully transformed output DataFrame.

    All checks produce WARNINGS, not errors -- the download is never blocked.
    The analyst reviews warnings and corrects source data as needed.

    Checks performed (in order):
      1. Empty output: no rows produced at all
      2. Missing required columns: any of ARCGIS_COLUMNS absent from output
      3. Blank unitid: ArcGIS requires a unit identifier on every row
      4. Duplicate unitid: ArcGIS expects unique unit identifiers per event
      5. Blank staffname: ArcGIS requires officer name on every row
      6. Blank unitshiftstart / unitshiftend: required for event timeline display
      7. Unknown staffrank: values outside APPROVED_RANKS fail ArcGIS domain validation

    The optional source_df argument is reserved for future use (e.g., cross-
    referencing output row numbers with source rows for precise warning messages).

    Returns:
        List of warning strings. Empty list means data is clean.
    """
    issues = []

    # Check 1: empty output
    if df.empty:
        issues.append("Output is empty — no rows were produced.")
        return issues  # remaining checks are meaningless on an empty DataFrame

    # Check 2: all 20 required ArcGIS columns present
    missing_cols = [c for c in ARCGIS_COLUMNS if c not in df.columns]
    if missing_cols:
        issues.append(f"Missing required output columns: {missing_cols}")

    # Check 3: blank unitid
    blank_unit = df["unitid"].apply(lambda v: not str(v).strip()).sum()
    if blank_unit:
        issues.append(f"{blank_unit} row(s) have blank unitid.")

    # Check 4: duplicate unitid values
    # Only check non-blank unitids -- blank unitid is already flagged above.
    # ArcGIS may overwrite or reject rows that share a unitid, so this is a
    # significant warning the analyst must resolve before importing.
    uid_series = df["unitid"].astype(str).str.strip()
    non_blank  = uid_series[uid_series != ""]
    dup_mask   = non_blank.duplicated(keep=False)  # mark ALL copies of each duplicate
    dup_count  = dup_mask.sum()
    dups       = non_blank[non_blank.duplicated()].unique().tolist()
    if dups:
        issues.append(
            f"⚠️ Duplicate unitid detected — {dup_count} row(s) share a unitid with another row. "
            f"ArcGIS may overwrite or reject duplicate units on import. "
            f"Repeated unitid value(s): {dups[:10]}"
        )

    # Check 5: blank staffname
    blank_name = df["staffname"].apply(lambda v: not str(v).strip()).sum()
    if blank_name:
        issues.append(f"{blank_name} row(s) have blank staffname.")

    # Check 6: blank shift datetimes
    for col in ("unitshiftstart", "unitshiftend"):
        if col in df.columns:
            blank = df[col].apply(lambda v: not str(v).strip()).sum()
            if blank:
                issues.append(f"{blank} row(s) have blank {col}.")

    # Check 7: staffrank values not in ArcGIS approved domain
    if "staffrank" in df.columns:
        # Build lowercase set for O(1) membership testing
        approved_lower = {r.lower() for r in APPROVED_RANKS}
        bad_ranks = df[
            df["staffrank"].apply(
                # Only flag non-blank values -- blank rank is a separate concern
                lambda v: bool(str(v).strip()) and str(v).strip().lower() not in approved_lower
            )
        ]["staffrank"].unique().tolist()
        if bad_ranks:
            issues.append(
                f"staffrank value(s) not in approved ArcGIS list: {bad_ranks}. "
                f"These rows are included but may fail ArcGIS validation."
            )

    return issues


# ===========================================================================
# TEMPLATE WRITING
# ===========================================================================

def write_to_template(
    output_df: pd.DataFrame,
    template_path: str,
) -> io.BytesIO:
    """
    Write the transformed output DataFrame into the ArcGIS batch upload Excel template.

    Four-phase process:
      Phase 1 -- Load the existing template workbook.
                 Using load_workbook() preserves ALL sheets (dropdowns, data
                 dictionary, column definitions) that ArcGIS requires in the
                 upload file. Writing to a new blank workbook would lose those sheets.
      Phase 2 -- Snapshot row-2 cell formatting from the "Staff List" sheet.
                 The template has sample data in row 2 with specific formatting
                 (font, fill, borders, alignment, number formats). We capture it
                 with copy.copy() so we can re-apply it to every new data row.
                 copy.copy() is required because openpyxl reuses style objects by
                 reference -- assigning the same object to multiple cells causes
                 unexpected style sharing and corruption.
      Phase 3 -- Clear existing data rows (rows 2 onward).
                 Null all cell values without deleting rows or the header row.
                 This removes sample data while preserving column widths, row
                 heights, and sheet-level formatting.
      Phase 4 -- Write output_df rows starting at row 2.
                 Python datetime objects are written as-is; openpyxl converts
                 them to the correct Excel serial number internally. Row-2
                 formatting is re-applied to each new cell.

    WHY RETURN BytesIO INSTEAD OF WRITING TO DISK:
    Streamlit's st.download_button() requires a bytes-like object. Writing to an
    in-memory BytesIO buffer avoids creating a temporary file on disk and works
    in read-only deployment environments.

    Args:
        output_df:     transformed DataFrame with ARCGIS_COLUMNS
        template_path: absolute path to Sample_Batch_Load_Event_Staff_Template.xlsx

    Returns:
        io.BytesIO positioned at byte 0, ready for Streamlit download.

    Raises:
        ValueError: if the template does not contain a "Staff List" sheet.
    """
    # Phase 1: load the template workbook preserving all sheets and styles
    wb = load_workbook(template_path)

    if "Staff List" not in wb.sheetnames:
        raise ValueError(
            "Template workbook does not contain a sheet named 'Staff List'. "
            f"Found sheets: {wb.sheetnames}"
        )

    ws = wb["Staff List"]

    # -----------------------------------------------------------------------
    # Phase 1b: build header -> column index map
    # -----------------------------------------------------------------------
    # Row 1 is the header. Map each lowercase header string to its 1-based
    # column index so we can locate the correct column for each ARCGIS_COLUMNS
    # field without assuming a fixed column order in the template.
    # -----------------------------------------------------------------------
    header_map = {}  # { "unitid": 1, "unitshift": 2, ... }
    for cell in ws[1]:
        if cell.value is not None:
            header_map[str(cell.value).strip().lower()] = cell.column

    # -----------------------------------------------------------------------
    # Phase 2: snapshot row-2 cell formatting
    # -----------------------------------------------------------------------
    row2_formats = {}  # { column_index: { style_attr: copied_value } }
    for cell in ws[2]:
        row2_formats[cell.column] = {
            "font":          copy.copy(cell.font),
            "fill":          copy.copy(cell.fill),
            "border":        copy.copy(cell.border),
            "alignment":     copy.copy(cell.alignment),
            "number_format": cell.number_format,  # string -- no copy needed
        }

    # -----------------------------------------------------------------------
    # Phase 3: clear existing data rows (row 2 onward)
    # -----------------------------------------------------------------------
    # Set all cell values to None (not "") -- None produces truly empty cells
    # in Excel; "" shows as an empty string which can interfere with ArcGIS
    # field parsing for numeric and coded domain fields.
    # -----------------------------------------------------------------------
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 2:
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).value = None

    # -----------------------------------------------------------------------
    # Phase 4: write output rows
    # -----------------------------------------------------------------------
    # These ArcGIS column names hold datetime values and need special treatment:
    # Python datetime object written + ArcGIS datetime number format applied.
    datetime_arcgis_cols = {"unitshiftstart", "unitshiftend"}

    # enumerate starting at 2 so row_idx maps directly to the Excel row number
    for row_idx, (_, data_row) in enumerate(output_df.iterrows(), start=2):
        for arcgis_col in ARCGIS_COLUMNS:
            # Look up which template column this ArcGIS field maps to
            col_idx = header_map.get(arcgis_col.lower())
            if col_idx is None:
                continue  # field not present in template header -- skip silently

            cell  = ws.cell(row=row_idx, column=col_idx)
            value = data_row.get(arcgis_col, "")

            # Write the cell value with type-appropriate handling
            if isinstance(value, datetime):
                # openpyxl converts Python datetime to Excel serial float internally.
                # Applying the number format causes Excel to display it as a date string.
                cell.value         = value
                cell.number_format = _OPENPYXL_DATE_FORMAT
            elif value == "" or (isinstance(value, float) and pd.isna(value)):
                # Explicit None for blank cells -- cleaner than empty string in Excel
                cell.value = None
            else:
                cell.value = value

            # Re-apply the row-2 formatting template to this cell
            fmt = row2_formats.get(col_idx)
            if fmt:
                if fmt["font"]:
                    cell.font = fmt["font"]
                if fmt["fill"]:
                    cell.fill = fmt["fill"]
                if fmt["border"]:
                    cell.border = fmt["border"]
                if fmt["alignment"]:
                    cell.alignment = fmt["alignment"]
                # Only override number_format for non-datetime cells AND only when
                # the template row had a custom format (not the default "General").
                # This prevents overwriting the ArcGIS datetime format set above.
                if (
                    arcgis_col not in datetime_arcgis_cols
                    and fmt["number_format"]
                    and fmt["number_format"] != "General"
                ):
                    cell.number_format = fmt["number_format"]

    # Save to in-memory buffer -- avoids temp files, works in all environments
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)  # rewind to byte 0 so caller can read or pass to st.download_button
    return buf
